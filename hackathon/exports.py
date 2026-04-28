from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import Hackathon, HackathonTeam, TeamMember


def _autosize_columns(ws):
    for idx, column_cells in enumerate(ws.columns, start=1):
        length = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = min(max(length + 2, 12), 48)


def build_participants_workbook(hackathon: Hackathon) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"
    headers = [
        "Хакатон",
        "ФИО",
        "Email",
        "Telegram ID",
        "Стек",
        "Команда",
        "Роль в команде",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    team_links = (
        HackathonTeam.objects.filter(hackathon=hackathon)
        .select_related("team", "team__captain")
        .prefetch_related("team__teammember_set", "team__teammember_set__user")
    )

    seen_user_ids = set()
    for link in team_links:
        team = link.team
        members = TeamMember.objects.filter(
            team=team,
            status=TeamMember.Status.ACCEPTED,
        ).select_related("user")

        for tm in members:
            uid = tm.user_id
            if uid in seen_user_ids:
                continue
            seen_user_ids.add(uid)
            u = tm.user
            role = "Капитан" if team.captain_id == u.id else "Участник"
            ws.append(
                [
                    hackathon.name,
                    u.full_name,
                    u.email,
                    u.telegram_id,
                    u.skills,
                    team.name,
                    role,
                ]
            )

    _autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_teams_workbook(hackathon: Hackathon) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Команды"
    headers = [
        "Хакатон",
        "Команда",
        "Капитан",
        "Telegram капитана",
        "Участников (принято)",
        "Стек",
        "Вакансии",
        "Набор открыт",
        "Дата подключения к хакатону",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    team_links = (
        HackathonTeam.objects.filter(hackathon=hackathon)
        .select_related("team", "team__captain")
    )

    for link in team_links.order_by("team__name"):
        team = link.team
        accepted = TeamMember.objects.filter(
            team=team,
            status=TeamMember.Status.ACCEPTED,
        ).count()
        cap = team.captain
        ws.append(
            [
                hackathon.name,
                team.name,
                cap.full_name,
                cap.telegram_id,
                accepted,
                team.tech_stack,
                team.vacancies,
                "да" if team.is_open else "нет",
                link.joined_at.isoformat() if link.joined_at else "",
            ]
        )

    _autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
